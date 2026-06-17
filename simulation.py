"""
怪物生成 Monte Carlo 模拟
=========================
模拟每张主题地图在每个刷怪波次中，场上怪物数量的期望值。

规则（源自 怪物生成.pdf）：
1. 每张地图有室内/室外独立的权值上限，累积达到上限后不再生成
2. 每局开始生成一次(第0波)，之后每 X 秒为一波，检查权值是否达上限
3. 每波最多生成一只怪物
4. 生成概率 = 基础概率 × 时间修正[波次] × 数量修正[当前场上数量]
5. 不满足条件（born_pos_type 不匹配、已达 max_num、已达 max_num_generated）则概率置0
6. 归一化后按概率随机选择怪物
7. 不建模死亡（纯累积，monster_weight 只加不减）
"""
import openpyxl
import random
import math
from collections import defaultdict
from dataclasses import dataclass, field
from typing import List, Dict, Tuple, Optional

# ============================================================
# 1. 数据结构定义
# ============================================================

@dataclass
class Monster:
    """怪物数据"""
    idx: int              # 在 map monster_class 中的索引 (0-7)
    name: str             # 怪物名称
    monster_weight: float           # 怪物权重（占用权值上限的量）
    time_weight: List[float]        # 时间修正数组 [波次] -> 权重
    num_weight: List[float]         # 数量修正数组 [场上数量] -> 权重
    max_num: int                    # 最大同时在场数
    max_num_generated: int          # 最大总生成数 (-1 表示无限)
    born_pos_type: int              # 1=室内, 2=室外, 3=两者


@dataclass
class MapConfig:
    """地图配置"""
    map_id: str
    region_id: str
    difficulty: str
    gen_prob: List[float]           # 8个怪物的基础生成概率
    outdoor_total_weight: float     # 室外权值上限
    indoor_total_weight: float      # 室内权值上限
    spawn_cycle: float              # 刷怪周期(秒)
    game_duration: float            # 游戏时长(秒)
    num_waves: int                  # 波次数（含第0波）

    @property
    def display_name(self) -> str:
        return f"{self.region_id}_{self.difficulty} ({self.map_id})"


@dataclass
class PoolState:
    """一个池子（室内或室外）的状态"""
    total_weight_limit: float
    current_total_weight: float = 0.0
    monster_counts: Dict[int, int] = field(default_factory=dict)   # monster_idx -> 当前场上数量
    monster_generated: Dict[int, int] = field(default_factory=dict) # monster_idx -> 总生成数

    def can_spawn(self) -> bool:
        """权值是否未达上限"""
        return self.current_total_weight < self.total_weight_limit


# ============================================================
# 2. 数据解析
# ============================================================

def parse_monster_table(filepath: str, monster_name_to_idx: Dict[str, int]) -> List[Monster]:
    """从 monster.xlsx 解析怪物数据，按 map monster_class 顺序返回"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['monster_table']

    # 找到数据起始行：跳过所有以 ## 开头的行
    data_start = 1
    for row_idx in range(1, min(ws.max_row + 1, 20)):
        cell_val = ws.cell(row=row_idx, column=1).value
        if cell_val is not None and str(cell_val).strip().startswith('##'):
            continue
        # 遇到第一个非注释行就是数据开始
        data_start = row_idx
        break

    name_to_monster = {}
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if row[1] is None:
            continue

        name = str(row[1]).strip()
        time_weight_str = str(row[25]) if row[25] is not None else ""
        num_weight_str = str(row[26]) if row[26] is not None else ""

        monster = Monster(
            idx=-1,  # 稍后映射
            name=name,
            monster_weight=float(row[24]) if row[24] is not None else 1.0,
            time_weight=_parse_float_list(time_weight_str),
            num_weight=_parse_float_list(num_weight_str),
            max_num=int(row[14]) if row[14] is not None else 99,
            max_num_generated=int(row[15]) if row[15] is not None else -1,
            born_pos_type=int(row[27]) if row[27] is not None else 1,
        )
        name_to_monster[name] = monster

    # 按 monster_name_to_idx 顺序映射
    result = []
    for name, idx in sorted(monster_name_to_idx.items(), key=lambda x: x[1]):
        if name in name_to_monster:
            m = name_to_monster[name]
            m.idx = idx
            result.append(m)
        else:
            # 可能名称不匹配，尝试模糊匹配
            found = False
            for n, m in name_to_monster.items():
                if name in n or n in name:
                    m.idx = idx
                    result.append(m)
                    found = True
                    break
            if not found:
                raise ValueError(f"怪物 '{name}' 未在 monster.xlsx 中找到！可用名称: {list(name_to_monster.keys())}")

    return result


def parse_map_table(filepath: str) -> Tuple[List[MapConfig], Dict[str, int], List[str]]:
    """从 map.xlsx 解析地图数据，返回 (地图列表, monster_name->idx映射, monster_class列表)"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['map_table']

    # 找到数据起始行：跳过所有以 ## 开头的行
    data_start = 1
    for row_idx in range(1, min(ws.max_row + 1, 20)):
        cell_val = ws.cell(row=row_idx, column=1).value
        if cell_val is not None and str(cell_val).strip().startswith('##'):
            continue
        data_start = row_idx
        break

    maps = []
    monster_name_to_idx = {}
    monster_class_list = []

    for row in ws.iter_rows(min_row=data_start, values_only=True):
        map_id = str(row[1]).strip() if row[1] else ""
        if not map_id or map_id == 'Hall':
            continue

        # 第一次遇到有效地图时解析 monster_class
        if not monster_class_list:
            monster_class_str = str(row[7]) if row[7] else ""
            if '|' in monster_class_str:
                monster_class_list = [s.strip() for s in monster_class_str.split('|')]
                for i, name in enumerate(monster_class_list):
                    monster_name_to_idx[name] = i

        # 解析 gen_prob
        gen_prob_str = str(row[8]) if row[8] is not None else ""
        gen_prob = _parse_float_list(gen_prob_str)
        if len(gen_prob) != len(monster_class_list):
            gen_prob = gen_prob[:len(monster_class_list)] + [0.0] * (len(monster_class_list) - len(gen_prob))

        spawn_cycle = float(row[11]) if row[11] is not None else 60
        game_duration = float(row[12]) if row[12] is not None else 900
        num_waves = int(game_duration / spawn_cycle) + 1  # 含第0波

        map_config = MapConfig(
            map_id=map_id,
            region_id=str(row[2]).strip() if row[2] else "",
            difficulty=str(row[3]).strip() if row[3] else "",
            gen_prob=gen_prob,
            outdoor_total_weight=float(row[9]) if row[9] is not None else 10,
            indoor_total_weight=float(row[10]) if row[10] is not None else 10,
            spawn_cycle=spawn_cycle,
            game_duration=game_duration,
            num_waves=num_waves,
        )
        maps.append(map_config)

    return maps, monster_name_to_idx, monster_class_list


def _parse_float_list(s: str) -> List[float]:
    """解析管道分隔的浮点数列表"""
    if not s or not s.strip():
        return []
    parts = s.strip().split('|')
    result = []
    for p in parts:
        p = p.strip()
        if p:
            try:
                result.append(float(p))
            except ValueError:
                result.append(0.0)
    return result


# ============================================================
# 3. 核心模拟逻辑
# ============================================================

def get_eligible_monsters(pool_type: str, monsters: List[Monster]) -> List[int]:
    """
    获取属于指定池子的怪物索引列表
    pool_type: 'outdoor'(室外, born_pos_type=2或3) 或 'indoor'(室内, born_pos_type=1或3)
    """
    if pool_type == 'outdoor':
        return [m.idx for m in monsters if m.born_pos_type in (2, 3)]
    else:
        return [m.idx for m in monsters if m.born_pos_type in (1, 3)]


def get_time_weight(monster: Monster, wave: int) -> float:
    """获取怪物在指定波次的时间修正权重，越界用最后一个值"""
    if not monster.time_weight:
        return 0.0
    if wave < len(monster.time_weight):
        return monster.time_weight[wave]
    return monster.time_weight[-1]


def get_num_weight(monster: Monster, current_count: int) -> float:
    """获取怪物在指定场上数量时的数量修正权重，越界用最后一个值"""
    if not monster.num_weight:
        return 0.0
    if current_count < len(monster.num_weight):
        return monster.num_weight[current_count]
    return monster.num_weight[-1]


def check_spawn_conditions(monster: Monster, pool_state: PoolState,
                           wave: int, gen_prob: float) -> bool:
    """检查怪物是否满足生成条件"""
    idx = monster.idx
    current_count = pool_state.monster_counts.get(idx, 0)
    total_generated = pool_state.monster_generated.get(idx, 0)

    # 基础概率为0
    if gen_prob <= 0:
        return False

    # 达到最大同时在场数
    if current_count >= monster.max_num:
        return False

    # 达到最大总生成数（-1 表示无限）
    if monster.max_num_generated >= 0 and total_generated >= monster.max_num_generated:
        return False

    # 时间修正为0
    if get_time_weight(monster, wave) <= 0:
        return False

    # 数量修正为0
    if get_num_weight(monster, current_count) <= 0:
        return False

    # 检查生成后权值是否会超出上限
    if pool_state.current_total_weight + monster.monster_weight > pool_state.total_weight_limit:
        return False

    return True


def compute_spawn_probs(map_cfg: MapConfig, monsters: List[Monster],
                        pool_state: PoolState, pool_type: str, wave: int) -> List[Tuple[int, float]]:
    """
    计算本波次各怪物的生成概率（未归一化），返回 [(monster_idx, raw_prob), ...]
    只会返回概率 > 0 的怪物
    """
    eligible_indices = get_eligible_monsters(pool_type, monsters)
    raw_probs = []

    for idx in eligible_indices:
        monster = monsters[idx]
        gen_prob = map_cfg.gen_prob[idx]

        if not check_spawn_conditions(monster, pool_state, wave, gen_prob):
            continue

        current_count = pool_state.monster_counts.get(idx, 0)
        tw = get_time_weight(monster, wave)
        nw = get_num_weight(monster, current_count)

        raw_prob = gen_prob * tw * nw
        if raw_prob > 0:
            raw_probs.append((idx, raw_prob))

    return raw_probs


def spawn_monster(monster_idx: int, monsters: List[Monster], pool_state: PoolState):
    """在池子中生成一只怪物，更新状态"""
    monster = monsters[monster_idx]
    pool_state.monster_counts[monster_idx] = pool_state.monster_counts.get(monster_idx, 0) + 1
    pool_state.monster_generated[monster_idx] = pool_state.monster_generated.get(monster_idx, 0) + 1
    pool_state.current_total_weight += monster.monster_weight


def simulate_one_game(map_cfg: MapConfig, monsters: List[Monster]) -> Dict:
    """
    模拟一局游戏，返回结果字典：
    {
        'outdoor': {wave: {monster_idx: count}},  # 每波结束时室外场上各怪物数量
        'indoor':  {wave: {monster_idx: count}},  # 每波结束时室内场上各怪物数量
    }
    """
    num_monsters = len(monsters)

    # 初始化室内外两个池子
    outdoor_state = PoolState(total_weight_limit=map_cfg.outdoor_total_weight)
    indoor_state = PoolState(total_weight_limit=map_cfg.indoor_total_weight)

    # 初始化计数
    for i in range(num_monsters):
        outdoor_state.monster_counts[i] = 0
        outdoor_state.monster_generated[i] = 0
        indoor_state.monster_counts[i] = 0
        indoor_state.monster_generated[i] = 0

    outdoor_result = {}
    indoor_result = {}

    for wave in range(map_cfg.num_waves):
        # --- 室外生成 ---
        if outdoor_state.can_spawn():
            probs = compute_spawn_probs(map_cfg, monsters, outdoor_state, 'outdoor', wave)
            if probs:
                total = sum(p for _, p in probs)
                if total > 0:
                    normalized = [(idx, p / total) for idx, p in probs]
                    chosen = _weighted_choice(normalized)
                    if chosen is not None:
                        spawn_monster(chosen, monsters, outdoor_state)

        # --- 室内生成 ---
        if indoor_state.can_spawn():
            probs = compute_spawn_probs(map_cfg, monsters, indoor_state, 'indoor', wave)
            if probs:
                total = sum(p for _, p in probs)
                if total > 0:
                    normalized = [(idx, p / total) for idx, p in probs]
                    chosen = _weighted_choice(normalized)
                    if chosen is not None:
                        spawn_monster(chosen, monsters, indoor_state)

        # 记录本波结束时的场上数量
        outdoor_result[wave] = dict(outdoor_state.monster_counts)
        indoor_result[wave] = dict(indoor_state.monster_counts)

    return {'outdoor': outdoor_result, 'indoor': indoor_result}


def _weighted_choice(weighted_items: List[Tuple[int, float]]) -> Optional[int]:
    """按权重随机选择一个索引"""
    if not weighted_items:
        return None
    r = random.random()
    cumulative = 0.0
    for idx, w in weighted_items:
        cumulative += w
        if r <= cumulative:
            return idx
    # 浮点精度兜底
    return weighted_items[-1][0]


# ============================================================
# 4. Monte Carlo 模拟
# ============================================================

def monte_carlo_simulate(map_cfg: MapConfig, monsters: List[Monster],
                         num_trials: int = 10000,
                         seed: int = 42) -> Dict:
    """
    对一张地图运行 Monte Carlo 模拟

    返回:
    {
        'map': MapConfig,
        'num_trials': int,
        'num_waves': int,
        'outdoor_expected': {wave: {monster_idx: expected_count}},
        'indoor_expected':  {wave: {monster_idx: expected_count}},
        'outdoor_total_expected': {wave: float},   # 每波室外场上期望总数
        'indoor_total_expected':  {wave: float},   # 每波室内场上期望总数
    }
    """
    random.seed(seed)
    num_monsters = len(monsters)
    num_waves = map_cfg.num_waves

    # 累加器：outdoor_accum[wave][idx] = 总计数
    outdoor_accum = {w: {i: 0.0 for i in range(num_monsters)} for w in range(num_waves)}
    indoor_accum = {w: {i: 0.0 for i in range(num_monsters)} for w in range(num_waves)}

    for trial in range(num_trials):
        result = simulate_one_game(map_cfg, monsters)

        for wave in range(num_waves):
            for idx in range(num_monsters):
                outdoor_accum[wave][idx] += result['outdoor'][wave].get(idx, 0)
                indoor_accum[wave][idx] += result['indoor'][wave].get(idx, 0)

        if (trial + 1) % 1000 == 0:
            print(f"  [{map_cfg.display_name}] 已完成 {trial + 1}/{num_trials} 次模拟...")

    # 计算期望
    outdoor_expected = {}
    indoor_expected = {}
    outdoor_total = {}
    indoor_total = {}

    for wave in range(num_waves):
        outdoor_expected[wave] = {}
        indoor_expected[wave] = {}
        outdoor_total[wave] = 0.0
        indoor_total[wave] = 0.0

        for idx in range(num_monsters):
            outdoor_exp = outdoor_accum[wave][idx] / num_trials
            indoor_exp = indoor_accum[wave][idx] / num_trials
            outdoor_expected[wave][idx] = round(outdoor_exp, 4)
            indoor_expected[wave][idx] = round(indoor_exp, 4)
            outdoor_total[wave] += outdoor_exp
            indoor_total[wave] += indoor_exp

        outdoor_total[wave] = round(outdoor_total[wave], 4)
        indoor_total[wave] = round(indoor_total[wave], 4)

    return {
        'map': map_cfg,
        'num_trials': num_trials,
        'num_waves': num_waves,
        'outdoor_expected': outdoor_expected,
        'indoor_expected': indoor_expected,
        'outdoor_total_expected': outdoor_total,
        'indoor_total_expected': indoor_total,
    }


# ============================================================
# 5. 结果输出
# ============================================================

def print_results(results: Dict, monsters: List[Monster]):
    """打印模拟结果到控制台"""
    map_cfg = results['map']
    print(f"\n{'='*80}")
    print(f"地图: {map_cfg.display_name}")
    print(f"  室外权值上限: {map_cfg.outdoor_total_weight}, 室内权值上限: {map_cfg.indoor_total_weight}")
    print(f"  刷怪周期: {map_cfg.spawn_cycle}s, 游戏时长: {map_cfg.game_duration}s, 总波次: {map_cfg.num_waves}")
    print(f"  模拟次数: {results['num_trials']}")
    print(f"  基础概率: {dict(zip([m.name for m in monsters], map_cfg.gen_prob))}")

    # 打印室内结果（更有趣，有7种怪物）
    print(f"\n  --- 室内期望 ---")
    header = f"{'波次':>5}"
    for m in monsters:
        if m.born_pos_type in (1, 3):
            header += f" {m.name[:4]:>8}"
    header += f" {'总计':>8}"
    print(f"  {header}")
    print(f"  {'-'* (len(header)-2)}")

    for wave in range(results['num_waves']):
        line = f"  {wave:>5}"
        for m in monsters:
            if m.born_pos_type in (1, 3):
                val = results['indoor_expected'][wave][m.idx]
                line += f" {val:>8.2f}"
        line += f" {results['indoor_total_expected'][wave]:>8.2f}"
        print(line)

    # 打印室外结果（只有废墟恶犬）
    print(f"\n  --- 室外期望（仅废墟恶犬）---")
    outdoor_monsters = [m for m in monsters if m.born_pos_type in (2, 3)]
    header = f"{'波次':>5}"
    for m in outdoor_monsters:
        header += f" {m.name[:4]:>8}"
    header += f" {'总计':>8}"
    print(f"  {header}")
    print(f"  {'-'* (len(header)-2)}")

    for wave in range(results['num_waves']):
        line = f"  {wave:>5}"
        for m in outdoor_monsters:
            val = results['outdoor_expected'][wave][m.idx]
            line += f" {val:>8.2f}"
        line += f" {results['outdoor_total_expected'][wave]:>8.2f}"
        print(line)


def export_to_excel(all_results: List[Dict], monsters: List[Monster],
                    output_path: str):
    """将所有结果导出到 Excel"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    # 删除默认 sheet
    wb.remove(wb.active)

    header_font = Font(bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='4472C4')
    data_font = Font(size=10)
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )
    center_align = Alignment(horizontal='center', vertical='center')

    indoor_monsters = [m for m in monsters if m.born_pos_type in (1, 3)]
    outdoor_monsters = [m for m in monsters if m.born_pos_type in (2, 3)]

    for results in all_results:
        map_cfg = results['map']
        sheet_name = map_cfg.display_name[:31]  # Excel sheet 名限制31字符

        ws = wb.create_sheet(title=sheet_name)

        # === 标题行 ===
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=20)
        title_cell = ws.cell(row=1, column=1, value=f"怪物生成模拟 - {map_cfg.display_name}")
        title_cell.font = Font(bold=True, size=14, color='1F4E79')

        # === 参数信息 ===
        info_data = [
            (f"地图ID: {map_cfg.map_id}", f"难度: {map_cfg.difficulty}"),
            (f"室外权值上限: {map_cfg.outdoor_total_weight}", f"室内权值上限: {map_cfg.indoor_total_weight}"),
            (f"刷怪周期: {map_cfg.spawn_cycle}s", f"游戏时长: {map_cfg.game_duration}s"),
            (f"总波次: {map_cfg.num_waves}", f"模拟次数: {results['num_trials']}"),
        ]
        for i, (left, right) in enumerate(info_data):
            ws.cell(row=3 + i, column=1, value=left).font = Font(size=10, bold=True)
            ws.cell(row=3 + i, column=3, value=right).font = Font(size=10)

        # === 基础概率行 ===
        base_row = 8
        ws.cell(row=base_row, column=1, value="基础概率:").font = Font(size=10, bold=True)
        for j, m in enumerate(monsters):
            ws.cell(row=base_row, column=2 + j, value=m.name).font = Font(size=9)
            ws.cell(row=base_row + 1, column=2 + j, value=map_cfg.gen_prob[m.idx]).font = Font(size=9)

        # === 怪物信息行 ===
        info_row = base_row + 3
        ws.cell(row=info_row, column=1, value="怪物属性:").font = Font(size=10, bold=True)
        headers_m = ["名称", "权重", "max_num", "max_gen", "出生点"]
        for j, h in enumerate(headers_m):
            ws.cell(row=info_row, column=2 + j, value=h).font = Font(size=9, bold=True)
        for i, m in enumerate(monsters):
            ws.cell(row=info_row + 1 + i, column=1, value=m.name).font = Font(size=9)
            ws.cell(row=info_row + 1 + i, column=2, value=m.monster_weight).font = Font(size=9)
            ws.cell(row=info_row + 1 + i, column=3, value=m.max_num).font = Font(size=9)
            ws.cell(row=info_row + 1 + i, column=4, value=m.max_num_generated).font = Font(size=9)
            type_map = {1: '室内', 2: '室外', 3: '两者'}
            ws.cell(row=info_row + 1 + i, column=5, value=type_map.get(m.born_pos_type, '?')).font = Font(size=9)

        # === 室内期望表 ===
        indoor_start_row = info_row + len(monsters) + 3
        _write_expectation_table(ws, indoor_start_row, "室内生成期望",
                                 results['indoor_expected'], results['indoor_total_expected'],
                                 indoor_monsters, results['num_waves'])

        # === 室外期望表 ===
        outdoor_start_row = indoor_start_row + results['num_waves'] + 5
        _write_expectation_table(ws, outdoor_start_row, "室外生成期望",
                                 results['outdoor_expected'], results['outdoor_total_expected'],
                                 outdoor_monsters, results['num_waves'])

        # 调整列宽
        for col_idx in range(1, 20):
            ws.column_dimensions[get_column_letter(col_idx)].width = 14

    wb.save(output_path)
    print(f"\n结果已导出至: {output_path}")


def _write_expectation_table(ws, start_row: int, title: str,
                              expected: Dict, total_expected: Dict,
                              monster_list: List[Monster], num_waves: int):
    """写入一张期望值表"""
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, size=10, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='4472C4')
    center_align = Alignment(horizontal='center', vertical='center')

    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(monster_list) + 2)
    ws.cell(row=start_row, column=1, value=title).font = Font(bold=True, size=12, color='1F4E79')

    header_row = start_row + 1
    ws.cell(row=header_row, column=1, value="波次").font = header_font
    ws.cell(row=header_row, column=1).fill = header_fill
    ws.cell(row=header_row, column=1).alignment = center_align

    for j, m in enumerate(monster_list):
        cell = ws.cell(row=header_row, column=2 + j, value=m.name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    total_col = len(monster_list) + 2
    cell = ws.cell(row=header_row, column=total_col, value="总计")
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align

    for wave in range(num_waves):
        row = header_row + 1 + wave
        ws.cell(row=row, column=1, value=wave).alignment = center_align
        for j, m in enumerate(monster_list):
            ws.cell(row=row, column=2 + j, value=expected[wave][m.idx]).alignment = center_align
        ws.cell(row=row, column=total_col, value=total_expected[wave]).alignment = center_align


# ============================================================
# 6. 主函数
# ============================================================

def main():
    map_path = r"E:\project\design\source\Datas\map.xlsx"
    monster_path = r"E:\project\design\source\Datas\monster.xlsx"
    output_path = r"C:\Users\wepie\PieBox\Projects\untitled-mpywndbonu7n\monster_simulation_results.xlsx"

    print("=" * 60)
    print("怪物生成 Monte Carlo 模拟")
    print("=" * 60)

    # 解析数据
    print("\n[1] 解析表格数据...")
    maps, name_to_idx, monster_class_list = parse_map_table(map_path)
    monsters = parse_monster_table(monster_path, name_to_idx)

    print(f"  解析到 {len(maps)} 张地图, {len(monsters)} 种怪物")
    print(f"  怪物列表: {[m.name for m in monsters]}")

    # 过滤需要模拟的主题地图（排除 Tutorial 和 Hall）
    theme_maps = [m for m in maps if m.region_id in ('Castle', 'Bloodmoon_Castle')]
    print(f"\n  主题地图: {len(theme_maps)} 张")
    for m in theme_maps:
        print(f"    - {m.display_name}: 室外权重上限={m.outdoor_total_weight}, "
              f"室内权重上限={m.indoor_total_weight}, 周期={m.spawn_cycle}s, "
              f"波次={m.num_waves}")

    # 运行模拟
    print(f"\n[2] 运行 Monte Carlo 模拟 (10000次/地图)...")
    all_results = []
    for map_cfg in theme_maps:
        print(f"\n  模拟 {map_cfg.display_name}...")
        result = monte_carlo_simulate(map_cfg, monsters, num_trials=10000)
        all_results.append(result)
        print_results(result, monsters)

    # 导出 Excel
    print(f"\n[3] 导出结果到 Excel...")
    export_to_excel(all_results, monsters, output_path)

    print(f"\n完成！")


if __name__ == '__main__':
    main()
